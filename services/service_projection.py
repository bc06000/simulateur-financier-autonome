from pathlib import Path

from modeles.simulation import Simulation
from services.service_stockage import ServiceStockage
from services.calcul_projection import CalculProjection


class ServiceProjection:

    _instance = None

    def __new__(cls):

        if cls._instance is None:

            cls._instance = super().__new__(cls)

            cls._instance.dernier_resultat = None
            cls._instance.simulations = []
            cls._instance.stockage = ServiceStockage()

            cls._instance._charger()

        return cls._instance

    def _charger(self):

        donnees = self.stockage.charger(
            "simulations.json"
        )

        self.simulations = []

        if not donnees:
            return

        for element in donnees:

            try:

                simulation = Simulation.from_dict(
                    element
                )

                self.simulations.append(
                    simulation
                )

            except Exception:

                pass

        if self.simulations:

            self.dernier_resultat = (
                self.simulations[-1]
            )

    def calculer(
        self,
        capital_initial,
        versement_mensuel,
        taux_annuel,
        duree,
    ):

        simulation = CalculProjection.calculer(
            capital_initial,
            versement_mensuel,
            taux_annuel,
            duree,
        )

        self.dernier_resultat = simulation

        return simulation

    def calculer_scenarios(
        self,
        capital_initial,
        versement_mensuel,
        taux_annuel,
        duree,
    ):

        scenarios = {}

        scenarios["prudent"] = CalculProjection.calculer(
            capital_initial,
            versement_mensuel,
            taux_annuel * 0.7,
            duree,
        )

        scenarios["normal"] = CalculProjection.calculer(
            capital_initial,
            versement_mensuel,
            taux_annuel,
            duree,
        )

        scenarios["optimiste"] = CalculProjection.calculer(
            capital_initial,
            versement_mensuel,
            taux_annuel * 1.3,
            duree,
        )

        return scenarios

    def _sauvegarder(self):

        donnees = [

            simulation.to_dict()

            for simulation in self.simulations

        ]

        self.stockage.sauvegarder(
            "simulations.json",
            donnees,
        )

        self._exporter_excel(donnees)

    def _exporter_excel(self, donnees):

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            fichier_excel = (
                Path.home()
                / "Desktop"
                / "Simulations_utilisateurs_complet.xlsx"
            )

            classeur = Workbook()

            feuille = classeur.active
            feuille.title = "Simulations"

            feuille.append([
                "N°", "ID", "Date", "Capital initial (€)",
                "Versement mensuel (€)", "Taux (%)", "Durée (ans)",
                "Total versé (€)", "Capital final (€)",
                "Plus-value (€)", "Performance (%)",
            ])

            for numero, simulation in enumerate(donnees, start=1):
                feuille.append([
                    numero,
                    simulation.get("id", ""),
                    simulation.get("date_creation", ""),
                    simulation.get("capital_initial", 0),
                    simulation.get("versement_mensuel", 0),
                    simulation.get("taux", 0),
                    simulation.get("duree", 0),
                    simulation.get("total_versements", 0),
                    simulation.get("capital_final", 0),
                    simulation.get("gains", 0),
                    simulation.get("performance", 0),
                ])

            historique = classeur.create_sheet("Historique mensuel")
            historique.append([
                "N° simulation", "ID", "Date simulation",
                "Mois", "Capital (€)", "Intérêts (€)",
            ])

            for numero, simulation in enumerate(donnees, start=1):
                for ligne in simulation.get("historique", []) or []:
                    historique.append([
                        numero,
                        simulation.get("id", ""),
                        simulation.get("date_creation", ""),
                        ligne.get("mois", ""),
                        ligne.get("capital", 0),
                        ligne.get("interets", 0),
                    ])

            for feuille_excel in (feuille, historique):
                for cellule in feuille_excel[1]:
                    cellule.font = Font(bold=True)

                feuille_excel.freeze_panes = "A2"
                feuille_excel.auto_filter.ref = feuille_excel.dimensions

                for colonne in feuille_excel.columns:
                    largeur = max(
                        len(str(cellule.value) if cellule.value is not None else "")
                        for cellule in colonne
                    )
                    lettre = get_column_letter(colonne[0].column)
                    feuille_excel.column_dimensions[lettre].width = min(
                        max(largeur + 2, 12), 35
                    )

            classeur.save(fichier_excel)
            print(f"[EXCEL] Mis à jour : {fichier_excel}")

        except Exception as erreur:
            print(f"[EXCEL] ERREUR : {erreur}")

    def ajouter_simulation(
        self,
        simulation,
    ):

        ids_existants = [
            element.id
            for element in self.simulations
            if isinstance(element.id, int)
            and element.id > 0
        ]

        simulation.id = (
            max(ids_existants) + 1
            if ids_existants
            else 1
        )

        self.simulations.append(
            simulation
        )

        self.dernier_resultat = simulation

        self._sauvegarder()

    def obtenir_simulations(self):

        return self.simulations

    def obtenir_simulation(
        self,
        index,
    ):

        if 0 <= index < len(self.simulations):

            return self.simulations[index]

        return None

    def supprimer_simulation(
        self,
        index,
    ):

        if 0 <= index < len(self.simulations):

            del self.simulations[index]

            if self.simulations:

                self.dernier_resultat = (
                    self.simulations[-1]
                )

            else:

                self.dernier_resultat = None

            self._sauvegarder()

    def vider(self):

        self.simulations.clear()

        self.dernier_resultat = None

        self._sauvegarder()

    def obtenir_dernier_resultat(self):

        return self.dernier_resultat
